#!/usr/bin/env python3
"""Extract Polish EGiB parcel/owner records from PDF or HTML into CSV/Excel.

Designed for reports containing repeated "INFORMACJA O DZIAŁCE" sections.
HTML is parsed directly and is preferred. PDF reports with broken/custom fonts
are processed with OCR.

Requirements:
  - beautifulsoup4
  - openpyxl for Excel output
  - PDF only: Poppler (pdfinfo, pdftoppm) and Tesseract with Polish language data

Installation:
  pip install beautifulsoup4 openpyxl
  # On macOS: brew install poppler tesseract tesseract-pol
  # On Ubuntu: sudo apt-get install poppler-utils tesseract-ocr tesseract-ocr-pol

Examples:
  python egib_extractor.py "Dane opisowe EGiB.html" --output parcels
  python egib_extractor.py Dane.pdf --output parcels --start-page 1 --end-page 20
  python egib_extractor.py report.html --output my_data --dpi 300
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional

try:
    from bs4 import BeautifulSoup, Tag
except ImportError:
    print("Error: beautifulsoup4 is required. Install with: pip install beautifulsoup4", file=sys.stderr)
    sys.exit(1)


PARCEL_ID_RE = re.compile(r"\b(\d{6}_\d[._]\d{4}[._]\d+(?:/\d+)?)\b")
PESEL_RE = re.compile(r"\bPESEL\s*:\s*(\d{11})\b", re.I)
REGON_RE = re.compile(r"\bREGON\s*:\s*(\d{9}|\d{14})\b", re.I)
SHARE_RE = re.compile(r"(?<!\d)(\d{1,4}\s*/\s*\d{1,4})(?!\d)")
# Areas in these Polish reports use a decimal comma.  Accepting a dot caused
# fragments of parcel IDs such as ``2.0004`` to be mistaken for hectares.
AREA_RE = re.compile(r"(?<!\d)(\d{1,5},\d{2,4})(?!\d)")
ROW_START_RE = re.compile(r"^\s*[|;]?\s*(\d{1,3})\s*[|:]?\s+(.*)$")


@dataclass
class Record:
    parcel_no: str
    parcel_id: str
    area_ha: str
    owner: str
    ownership_type: str
    share: str
    pesel: str
    regon: str
    register_unit: str
    register_group: str
    voivodeship: str
    county: str
    cadastral_unit: str
    precinct_name: str
    precinct_no: str
    source_pages: str
    confidence: str
    warnings: str


def run(cmd: list[str], *, quiet: bool = False) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        check=True,
        text=True,
        stdout=subprocess.DEVNULL if quiet else subprocess.PIPE,
        stderr=subprocess.DEVNULL if quiet else subprocess.PIPE,
    )


def require_command(name: str, install_hint: Optional[str] = None) -> None:
    if not shutil.which(name):
        hint = f"\n  Install hint: {install_hint}" if install_hint else ""
        raise SystemExit(f"❌ Required command not found: {name}{hint}")


def page_count(pdf: Path) -> int:
    result = run(["pdfinfo", str(pdf)])
    match = re.search(r"^Pages:\s+(\d+)", result.stdout, re.M)
    if not match:
        raise RuntimeError("Could not determine PDF page count")
    return int(match.group(1))


def clean(text: str) -> str:
    text = unicodedata.normalize("NFC", text).replace("\u00ad", "")
    text = text.replace("WŁADYStLAWA", "WŁADYSŁAWA")
    return "\n".join(re.sub(r"[ \t]+", " ", line).strip() for line in text.splitlines())


def ocr_pages(
    pdf: Path, work: Path, start: int, end: int, dpi: int, lang: str, psm: int
) -> dict[int, str]:
    print(f"📄 Converting PDF pages {start}-{end} to images at {dpi} DPI...", file=sys.stderr)
    images = work / "images"
    texts = work / "text"
    images.mkdir(parents=True, exist_ok=True)
    texts.mkdir(parents=True, exist_ok=True)
    run([
        "pdftoppm", "-f", str(start), "-l", str(end), "-r", str(dpi),
        "-png", str(pdf), str(images / "page")
    ], quiet=True)

    print(f"🔍 Running OCR with Tesseract (language: {lang})...", file=sys.stderr)
    result: dict[int, str] = {}
    for offset, image in enumerate(sorted(images.glob("page-*.png"))):
        page = start + offset
        base = texts / f"page-{page:04d}"
        run(["tesseract", str(image), str(base), "-l", lang, "--psm", str(psm)], quiet=True)
        result[page] = clean(base.with_suffix(".txt").read_text(encoding="utf-8", errors="replace"))
        print(f"  ✓ OCR page {page}/{end}", file=sys.stderr)
    return result


def fuzzy_label_value(text: str, labels: list[str]) -> str:
    for label in labels:
        # Horizontal whitespace only: ``\s`` also consumes newlines and used
        # to attach the following label/value to an empty field.
        match = re.search(label + r"[ \t]*:[ \t]*([^\n|]+)", text, re.I)
        if match:
            return match.group(1).strip()
    return ""


def normalize_context(records: list[Record]) -> None:
    """Validate repeated report metadata and repair it by document consensus."""
    if not records:
        return

    def mode(values: list[str]) -> str:
        return Counter(value for value in values if value).most_common(1)[0][0] if any(values) else ""

    voivodeship = mode([
        r.voivodeship for r in records
        if re.fullmatch(r"[A-Za-zĄĆĘŁŃÓŚŹŻąćęłńóśźż -]+", r.voivodeship)
        and not re.search(r"Powiat|Jednostka|obr[eę]b", r.voivodeship, re.I)
    ])
    county = mode([
        r.county for r in records
        if re.fullmatch(r"[A-Za-zĄĆĘŁŃÓŚŹŻąćęłńóśźż -]+", r.county)
        and not re.search(r"Jednostka|obr[eę]b", r.county, re.I)
    ])
    cadastral_unit = mode([
        r.cadastral_unit for r in records
        if re.fullmatch(r"[A-Za-zĄĆĘŁŃÓŚŹŻąćęłńóśźż -]+", r.cadastral_unit)
        and not re.search(r"Nazwa|obr[eę]b|Wojew[oó]dztwo", r.cadastral_unit, re.I)
    ])

    names_by_precinct: dict[str, list[str]] = defaultdict(list)
    for record in records:
        precinct = record.parcel_id.split(".")[1]
        if (
            re.fullmatch(r"[A-Za-zĄĆĘŁŃÓŚŹŻąćęłńóśźż -]+", record.precinct_name)
            and not re.search(r"Numer|Informacja|obr[eę]b", record.precinct_name, re.I)
        ):
            names_by_precinct[precinct].append(record.precinct_name)
    precinct_names = {key: mode(values) for key, values in names_by_precinct.items()}

    for record in records:
        precinct = record.parcel_id.split(".")[1]
        record.precinct_no = precinct
        record.precinct_name = precinct_names.get(precinct, "")
        record.voivodeship = voivodeship
        record.county = county
        record.cadastral_unit = cadastral_unit
        if not re.fullmatch(r"G\.\d+", record.register_unit):
            record.register_unit = ""
        if not re.match(r"^\d+\s*\(", record.register_group):
            record.register_group = ""


def direct_cells(row: Tag) -> list[Tag]:
    return row.find_all(["td", "th"], recursive=False)


def table_rows(table: Tag) -> list[Tag]:
    """Return rows belonging to this table, excluding rows of nested tables."""
    rows: list[Tag] = []
    for row in table.find_all("tr"):
        if row.find_parent("table") is table:
            rows.append(row)
    return rows


def html_text(node: Tag) -> str:
    return clean(node.get_text(" ", strip=True))


def html_label_value(section: Tag, label: str) -> str:
    for row in section.find_all("tr"):
        cells = direct_cells(row)
        if len(cells) >= 2 and re.search(label, html_text(cells[0]), re.I):
            return html_text(cells[1])
    return ""


def extract_html(html_path: Path) -> list[Record]:
    raw = html_path.read_bytes()
    # The downloaded Geoportal report declares Windows-1250. Respect a
    # declared charset, with safe fallbacks for exported variants.
    header = raw[:4096].decode("ascii", errors="ignore")
    declared = re.search(r"charset\s*=\s*['\"]?([\w-]+)", header, re.I)
    encodings = [declared.group(1) if declared else "", "cp1250", "utf-8"]
    text = ""
    for encoding in filter(None, encodings):
        try:
            text = raw.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if not text:
        text = raw.decode("cp1250", errors="replace")

    soup = BeautifulSoup(text, "html.parser")
    sections = soup.select("div.dzialka_info")
    records: list[Record] = []
    for index, section in enumerate(sections, 1):
        section_text = html_text(section)
        parcel_match = PARCEL_ID_RE.search(section_text)
        if not parcel_match:
            continue
        parcel_id = parcel_match.group(1)
        parcel_no = parcel_id.rsplit(".", 1)[-1]

        owners: list[dict[str, str]] = []
        area = ""
        for table in section.find_all("table"):
            rows = table_rows(table)
            if not rows:
                continue
            header_text = " ".join(html_text(cell) for cell in direct_cells(rows[0]))
            if re.search(r"Podmiot\s+ewidencyjny", header_text, re.I):
                for row in rows[1:]:
                    cells = direct_cells(row)
                    if len(cells) < 2 or not re.fullmatch(r"\d+", html_text(cells[0])):
                        continue
                    owner_raw = html_text(cells[1])
                    pesel_match = PESEL_RE.search(owner_raw)
                    regon_match = REGON_RE.search(owner_raw)
                    owner = clean(re.sub(r"\b(?:PESEL|REGON)\s*:\s*\d+\b", "", owner_raw, flags=re.I))
                    owners.append({
                        "owner": owner,
                        "ownership_type": html_text(cells[2]) if len(cells) > 2 else "",
                        "share": html_text(cells[3]).replace(" ", "") if len(cells) > 3 else "",
                        "pesel": pesel_match.group(1) if pesel_match else "",
                        "regon": regon_match.group(1) if regon_match else "",
                    })
            if re.search(r"Nr\s+dzia[łl]ki", header_text, re.I):
                for row in rows[1:]:
                    cells = direct_cells(row)
                    if cells and PARCEL_ID_RE.search(html_text(cells[0])):
                        area = html_text(cells[3]).replace(",", ".") if len(cells) > 3 else ""
                        break

        if not owners:
            owners = [{"owner": "", "ownership_type": "", "share": "", "pesel": "", "regon": ""}]
        context = {
            "register_unit": html_label_value(section, r"Jednostka\s+rejestrowa"),
            "register_group": html_label_value(section, r"Grupa\s+rejestrowa"),
            "voivodeship": html_label_value(section, r"Wojew[oó]dztwo"),
            "county": html_label_value(section, r"Powiat"),
            "cadastral_unit": html_label_value(section, r"Jednostka\s+ewidencyjna"),
            "precinct_name": html_label_value(section, r"Nazwa\s+obr[eę]bu"),
            "precinct_no": html_label_value(section, r"Numer\s+obr[eę]bu"),
        }
        for owner in owners:
            warnings = [
                warning for missing, warning in (
                    (not area, "missing_area"),
                    (not owner["owner"], "missing_owner"),
                    (not owner["share"], "missing_share"),
                    (not owner["ownership_type"], "missing_ownership_type"),
                ) if missing
            ]
            confidence = "high" if not warnings else ("medium" if len(warnings) <= 2 else "low")
            records.append(Record(
                parcel_no=parcel_no, parcel_id=parcel_id, area_ha=area,
                source_pages=f"html:{index}", confidence=confidence,
                warnings=";".join(warnings), **owner, **context,
            ))
    normalize_context(records)
    return records


def split_sections(pages: dict[int, str]) -> list[tuple[list[int], str]]:
    """Split records and attach a page-start continuation to the prior record."""
    sections: list[tuple[list[int], str]] = []
    current_pages: list[int] = []
    current = ""
    heading = re.compile(r"INFORMACJA\s+O\s+DZI[AĄ][ŁL]CE", re.I)
    for page, text in pages.items():
        pieces = heading.split(text)
        prefix = pieces[0]
        if current and prefix.strip():
            current += "\n" + prefix
            if page not in current_pages:
                current_pages.append(page)
        for piece in pieces[1:]:
            if current.strip():
                sections.append((current_pages, current))
            current = "INFORMACJA O DZIAŁCE\n" + piece
            current_pages = [page]
    if current.strip():
        sections.append((current_pages, current))
    return sections


def extract_owner_entries(owner_text: str) -> list[dict[str, str]]:
    # Discard the section metadata and keep only the owner table body.
    owner_lines = owner_text.splitlines()
    header_indexes = [i for i, line in enumerate(owner_lines) if re.search(r"Podmiot\s+ewidencyjny", line, re.I)]
    if header_indexes:
        owner_text = "\n".join(owner_lines[header_indexes[-1] + 1 :])
    lines = [line.strip(" |;") for line in owner_text.splitlines() if line.strip(" |;")]
    lines = [line for line in lines if not re.search(r"Lp\s+.*Podmiot|Charakter.*Udział", line, re.I)]
    entries: list[list[str]] = []
    pending_name: list[str] = []
    current: list[str] | None = None

    for line in lines:
        # Person names are printed immediately above their numbered address row.
        if re.search(r"\bRodzice\s*:", line, re.I):
            if current:
                entries.append(current)
                current = None
            pending_name = [line]
            continue
        row = ROW_START_RE.match(line)
        # Avoid treating postal codes, dates, and identifiers as row numbers.
        if row and int(row.group(1)) <= 99 and not re.match(r"\d{2}-\d{3}", line):
            if current:
                entries.append(current)
            current = pending_name + [row.group(2)]
            pending_name = []
        elif current is not None:
            current_joined = " ".join(current)
            complete = re.search(
                r"(?:Własność|Współwłasność|Użytkowanie wieczyste|Inny rodzaj władania|Zarząd|Trwały zarząd)",
                current_joined, re.I,
            ) and (SHARE_RE.search(current_joined) or re.search(r"(?:^|\s)11\s*$", current_joined))
            starts_new_owner = re.search(
                r"\bRodzice\s*:|\b(?:URZ[AĄ]D|GMINA|SKARB|PA[NŃ]STWA|SP[ÓO][ŁL]KA|PARAFIA)\b",
                line, re.I,
            )
            if complete and starts_new_owner:
                entries.append(current)
                current = None
                pending_name = [line]
            else:
                current.append(line)
        else:
            pending_name.append(line)
    if current:
        entries.append(current)
    elif pending_name:
        entries.append(pending_name)

    parsed: list[dict[str, str]] = []
    for entry in entries:
        # Remove OCR fragments of the parcel-table header which sometimes sit
        # on the final owner row because the two tables touch vertically.
        entry = [
            line for line in entry
            if not re.search(r"Po[lł]o[zż]enie|Klasou[zż]ytki|Pow\.\s*(?:dzia|$)|Nr\s*dz[ił]a", line, re.I)
        ]
        joined = " ".join(entry)
        share_matches = list(SHARE_RE.finditer(joined))
        ownership = ""
        ownership_match = re.search(
            r"(Własność|Współwłasność|Użytkowanie wieczyste|Inny rodzaj władania|Zarząd|Trwały zarząd)",
            joined, re.I,
        )
        if ownership_match:
            ownership = ownership_match.group(1)
        # Broken word segmentation frequently turns Własność into two pieces
        # around the address: "WŁ ... ASNOŚĆ".
        fuzzy_start = re.search(r"\bW[ŁLTI]\b", joined, re.I)
        fuzzy_ownership = fuzzy_start and re.search(
            r"\b(?:A[SŚ]NO[SŚ][ĆC]|JASNOSE)\b", joined, re.I
        )
        if not ownership and fuzzy_ownership:
            ownership = "Własność"
        # Prefer the first fraction after the ownership label. This avoids
        # taking apartment numbers (2/58, 35/30, 7/9) as ownership shares.
        anchor = ownership_match.end() if ownership_match else (fuzzy_start.end() if fuzzy_start else 0)
        share_match = next((m for m in share_matches if m.start() >= anchor), None)
        raw_share = share_match.group(1).replace(" ", "") if share_match else ""
        share = raw_share
        fuzzy_share = re.search(
            r"\bW[ŁLTI]\b(?:\s+[\"'A-Z]{1,2}){0,2}\s*(1?1|41|4M|4/[12])\b",
            joined, re.I,
        )
        if not share and fuzzy_share:
            token = fuzzy_share.group(1).upper()
            share = "1/2" if token == "4/2" else "1/1"
        if not share and ownership_match:
            after_ownership = joined[ownership_match.end() : ownership_match.end() + 20]
            if re.search(r"(?:^|\s)(?:11|41|4M)\b", after_ownership, re.I):
                share = "1/1"
        if share:
            numerator, denominator = map(int, share.split("/"))
            if numerator == 4 and denominator <= 2:
                share = f"1/{denominator}"
        # In narrow rightmost table cells OCR commonly drops the slash from
        # 1/1 and returns a trailing "11".  Only apply this correction when
        # an ownership label is present, avoiding unrelated address numbers.
        if not share and ownership and re.search(r"(?:^|\s)(?:11|41|4M)\s*$", joined, re.I):
            share = "1/1"

        pesel = (PESEL_RE.search(joined).group(1) if PESEL_RE.search(joined) else "")
        regon = (REGON_RE.search(joined).group(1) if REGON_RE.search(joined) else "")
        owner_parts: list[str] = []
        for line in entry:
            if re.search(r"\b(?:PESEL|REGON)\s*:", line, re.I):
                continue
            # Remove only the final slash value (the share), preserving address
            # fractions such as 2/58.
            if raw_share:
                matches = list(SHARE_RE.finditer(line))
                target = next((m for m in matches if m.group(0).replace(" ", "") == raw_share), None)
                if target:
                    line = line[:target.start()] + line[target.end():]
            if share == "1/1":
                line = re.sub(r"(?:^|\s)(?:11|41|4M)\s*$", "", line, flags=re.I)
            line = re.sub(r"\b(?:Własność|Współwłasność|Użytkowanie wieczyste|Inny rodzaj władania|Zarząd|Trwały zarząd)\b", "", line, flags=re.I)
            line = re.sub(r"^\s*\d{1,3}\s*[|:]?\s*", "", line)
            line = re.sub(r"\bRodzice\s*:[^;]+?\s+A\s*$", lambda m: m.group(0)[:-2], line, flags=re.I)
            line = re.sub(
                r"\bW[ŁLTI]\b(?:\s+[\"'A-Z]{1,2}){0,2}\s*(?:11|41|4M|4/[12])?",
                "", line, flags=re.I,
            )
            line = re.sub(r"\b(?:A[SŚ]NO[SŚ][ĆC]|JASNOSE)\b", "", line, flags=re.I)
            owner_parts.append(line.strip(" |;"))
        owner = re.sub(r"\s+", " ", " ".join(filter(None, owner_parts))).strip()
        bad_owner = re.search(
            r"Po[lł]o[zż]enie|Klasou[zż]ytki|Nr\s*dz[ił]a|Oznaczenie\s*[|:]\s*Pow|Informacja\s+o\s+budynkach",
            owner, re.I,
        )
        if owner and not bad_owner and (" " in owner or re.search(r"\b(?:GMINA|URZ[AĄ]D|SKARB|NIEUSTALONY)\b", owner, re.I)):
            parsed.append({"owner": owner, "ownership_type": ownership, "share": share, "pesel": pesel, "regon": regon})
    return parsed


def parse_section(pages: list[int], text: str) -> list[Record]:
    # Search the original OCR text. Removing every space concatenates the
    # slash suffix with the following land-use word ("190/1 RIVa" ->
    # "190/1RIVa"), making the regex backtrack to the incorrect ID "190".
    parcel_matches = list(PARCEL_ID_RE.finditer(text))
    if not parcel_matches:
        return []
    table_pos = re.search(r"Nr\s+dz[ił]a[łt]ki", text, re.I)
    owner_text = text[:table_pos.start()] if table_pos else text[: parcel_matches[0].start()]
    parcel_text = text[table_pos.start():] if table_pos else text
    # The parcel ID is the first cadastral ID in a parcel section. Later IDs
    # can belong to buildings and must not replace it (especially when the
    # parcel has a slash suffix such as 190/1).
    parcel_id = parcel_matches[0].group(1)
    parcel_no = parcel_id.rsplit(".", 1)[-1]
    # The full cadastral ID is more reliable than the narrow parcel-number
    # column.  In that column OCR often drops a slash (173/1 -> 1731) or joins
    # the first area digit (125 -> 1250), so never overwrite a valid full ID.
    areas = AREA_RE.findall(parcel_text)
    area = areas[0].replace(",", ".") if areas else ""
    owners = extract_owner_entries(owner_text)
    if not owners:
        owners = [{"owner": "", "ownership_type": "", "share": "", "pesel": "", "regon": ""}]

    context = {
        "register_unit": fuzzy_label_value(text, [r"Jednostka\s+rejestrowa"]),
        "register_group": fuzzy_label_value(text, [r"Grupa\s+rejestrowa"]),
        "voivodeship": fuzzy_label_value(text, [r"Wojew[oó]dztwo"]),
        "county": fuzzy_label_value(text, [r"Powiat"]),
        "cadastral_unit": fuzzy_label_value(text, [r"Jednostka\s+ewidencyjna"]),
        "precinct_name": fuzzy_label_value(text, [r"Nazwa\s+obr[eę]bu"]),
        "precinct_no": fuzzy_label_value(text, [r"Numer\s+obr[eę]bu"]),
    }
    result: list[Record] = []
    for owner in owners:
        warnings: list[str] = []
        if not area:
            warnings.append("missing_area")
        if not owner["owner"]:
            warnings.append("missing_owner")
        if not owner["share"]:
            warnings.append("missing_share")
        if not owner["ownership_type"]:
            warnings.append("missing_ownership_type")
        confidence = "high" if not warnings else ("medium" if len(warnings) <= 2 else "low")
        result.append(Record(
            parcel_no=parcel_no, parcel_id=parcel_id, area_ha=area,
            source_pages=",".join(map(str, pages)), confidence=confidence,
            warnings=";".join(warnings), **owner, **context,
        ))
    return result


def write_csv(records: list[Record], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(Record.__dataclass_fields__)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter=";")
        writer.writeheader()
        writer.writerows(asdict(record) for record in records)


def write_xlsx(records: list[Record], path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "Parcels and owners"
    headers = list(Record.__dataclass_fields__)
    ws.append(headers)
    for record in records:
        ws.append([getattr(record, key) for key in headers])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {1: 13, 2: 25, 3: 12, 4: 52, 5: 24, 6: 10, 7: 15, 8: 16}
    for index in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(index, 20)
    for cell in ws["C"][1:]:
        cell.number_format = "0.0000"
    ws.sheet_view.showGridLines = False
    wb.save(path)
    return True


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Output files:
  <output>.csv      - Semicolon-separated CSV with UTF-8 BOM
  <output>.xlsx     - Formatted Excel spreadsheet
  <output>_summary.json - Processing summary with statistics

For PDF processing, ensure Poppler and Tesseract are installed:
  macOS: brew install poppler tesseract tesseract-pol
  Ubuntu: sudo apt-get install poppler-utils tesseract-ocr tesseract-ocr-pol
"""
    )
    parser.add_argument("input", type=Path, help="Geoportal report in PDF or HTML format")
    parser.add_argument("--output", "-o", type=Path, default=Path("parcels"), 
                        help="Output basename (default: parcels)")
    parser.add_argument("--start-page", type=int, default=1, 
                        help="First page to process (PDF only, default: 1)")
    parser.add_argument("--end-page", type=int, 
                        help="Last page to process (PDF only, default: all)")
    parser.add_argument("--dpi", type=int, default=200, 
                        help="DPI for PDF rendering (higher = slower but more accurate, default: 200)")
    parser.add_argument("--lang", default="pol+eng", 
                        help="OCR language(s), e.g. 'pol' or 'pol+eng' (default: pol+eng)")
    parser.add_argument(
        "--psm", type=int, default=3,
        help="Tesseract page segmentation mode (default: 3 = auto, 6 = single column)",
    )
    parser.add_argument("--keep-ocr", action="store_true", 
                        help="Keep OCR text files in <output>_ocr/ directory")
    parser.add_argument("--version", action="version", version="EGiB Extractor 1.0")
    args = parser.parse_args()

    if not args.input.is_file():
        raise SystemExit(f"❌ Input file not found: {args.input}")
    
    suffix = args.input.suffix.lower()
    print(f"📂 Processing: {args.input} ({suffix.lstrip('.')})", file=sys.stderr)
    
    pages_summary: list[int] | None = None
    temp_context: tempfile.TemporaryDirectory[str] | None = None
    work: Path | None = None
    
    if suffix in {".html", ".htm"}:
        print("🔍 Parsing HTML file...", file=sys.stderr)
        records = extract_html(args.input)
    elif suffix == ".pdf":
        for command, hint in [
            ("pdfinfo", "brew install poppler / sudo apt-get install poppler-utils"),
            ("pdftoppm", "brew install poppler / sudo apt-get install poppler-utils"),
            ("tesseract", "brew install tesseract / sudo apt-get install tesseract-ocr")
        ]:
            require_command(command, hint)
        total = page_count(args.input)
        print(f"📊 PDF has {total} pages", file=sys.stderr)
        end = min(args.end_page or total, total)
        if not (1 <= args.start_page <= end):
            raise SystemExit("❌ Invalid page range")
        pages_summary = [args.start_page, end]
        temp_context = tempfile.TemporaryDirectory(prefix="parcel_ocr_")
        work = Path(temp_context.name)
        pages = ocr_pages(args.input, work, args.start_page, end, args.dpi, args.lang, args.psm)
        print("📝 Extracting parcel and owner records...", file=sys.stderr)
        records = [record for page_list, section in split_sections(pages) for record in parse_section(page_list, section)]
        normalize_context(records)
    else:
        raise SystemExit("❌ Unsupported input format; use .pdf, .html, or .htm")
    
    print(f"✅ Found {len(records)} records", file=sys.stderr)
    records.sort(key=lambda r: (r.parcel_id, r.owner))

    csv_path = args.output.with_suffix(".csv")
    xlsx_path = args.output.with_suffix(".xlsx")
    json_path = args.output.with_name(args.output.name + "_summary.json")
    
    print(f"💾 Writing output files...", file=sys.stderr)
    write_csv(records, csv_path)
    print(f"  ✓ CSV: {csv_path}", file=sys.stderr)
    
    xlsx_written = write_xlsx(records, xlsx_path)
    if xlsx_written:
        print(f"  ✓ Excel: {xlsx_path}", file=sys.stderr)
    else:
        print(f"  ⚠ Excel skipped (openpyxl not installed)", file=sys.stderr)
    
    unique_parcels = len({r.parcel_id for r in records})
    warning_rows = sum(bool(r.warnings) for r in records)
    
    summary = {
        "source": str(args.input), 
        "format": suffix.lstrip("."), 
        "pages": pages_summary,
        "records": len(records), 
        "parcels": unique_parcels,
        "warning_rows": warning_rows,
        "csv": str(csv_path), 
        "xlsx": str(xlsx_path) if xlsx_written else None,
    }
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  ✓ Summary: {json_path}", file=sys.stderr)
    
    if args.keep_ocr and suffix == ".pdf" and work is not None:
        target = args.output.with_name(args.output.name + "_ocr")
        if target.exists():
            shutil.rmtree(target)
        shutil.copytree(work / "text", target)
        print(f"  ✓ OCR text saved to: {target}", file=sys.stderr)
    
    if temp_context is not None:
        temp_context.cleanup()
    
    print(f"\n✨ Done! Extracted {len(records)} records from {unique_parcels} parcels.", file=sys.stderr)
    if warning_rows > 0:
        print(f"⚠️  {warning_rows} records have warnings (check 'warnings' column in CSV)", file=sys.stderr)
    
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
